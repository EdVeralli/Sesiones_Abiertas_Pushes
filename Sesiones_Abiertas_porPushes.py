# -*- coding: utf-8 -*-
"""

aws-azure-login --configure --profile default

aws-azure-login --profile default --mode=gui



Script para ejecutar query de Sesiones Abiertas por Pushes (starting_cause) en Athena
Guarda resultado en CSV y escribe el resultado en la celda D4 de Excel
Lee configuracion de fechas desde archivo config_fechas.txt

MODOS SOPORTADOS:
1. MES COMPLETO: Especificar MES y AÑO (comportamiento original)
2. RANGO PERSONALIZADO: Especificar FECHA_INICIO y FECHA_FIN

IMPORTANTE: El Excel SIEMPRE se crea NUEVO desde cero con estructura de Dashboard
Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBAConsumeBoti
"""
import boto3
import awswrangler as wr
import pandas as pd
from datetime import datetime
from calendar import monthrange
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': 'config_fechas.txt'
}

# ==================== FUNCIONES ====================

def read_date_config(config_file):
    """
    Lee el archivo de configuracion y determina el modo:
    - MODO 1: MES + AÑO (mes completo)
    - MODO 2: FECHA_INICIO + FECHA_FIN (rango personalizado)
    
    Retorna: (modo, fecha_inicio, fecha_fin, mes, anio, descripcion)
    """
    try:
        if not os.path.exists(config_file):
            print("[ERROR] No se encuentra el archivo: {}".format(config_file))
            print("    Creando archivo de ejemplo...")
            
            with open(config_file, 'w', encoding='utf-8') as f:
                f.write("# ========================================\n")
                f.write("# Configuracion de fechas para el reporte\n")
                f.write("# ========================================\n\n")
                f.write("# MODO 1: Mes completo (comportamiento original)\n")
                f.write("# Descomentar estas lineas para usar mes completo:\n")
                f.write("MES=10\n")
                f.write("AÑO=2025\n\n")
                f.write("# MODO 2: Rango de fechas personalizado\n")
                f.write("# Descomentar estas lineas para usar rango personalizado:\n")
                f.write("# FECHA_INICIO=2025-10-01\n")
                f.write("# FECHA_FIN=2025-10-15\n\n")
                f.write("# NOTA: Si ambos modos estan configurados, se usa el MODO 2 (rango personalizado)\n")
            
            print("    Archivo creado: {}".format(config_file))
            return 'mes', None, None, 10, 2025, "octubre 2025"
        
        # Leer el archivo y buscar ambos modos
        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                # MODO 1: MES y AÑO
                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)
                
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)
                
                # MODO 2: FECHA_INICIO y FECHA_FIN
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()
                
                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()
        
        # PRIORIDAD: Si hay FECHA_INICIO y FECHA_FIN, usar MODO 2 (rango personalizado)
        if fecha_inicio_str and fecha_fin_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
                
                if fecha_inicio > fecha_fin:
                    print("[ERROR] FECHA_INICIO no puede ser posterior a FECHA_FIN")
                    return None, None, None, None, None, None
                
                # Descripcion para el rango
                descripcion = "{} al {}".format(
                    fecha_inicio.strftime('%d/%m/%Y'),
                    fecha_fin.strftime('%d/%m/%Y')
                )
                
                print("[INFO] Modo: RANGO PERSONALIZADO")
                return 'rango', fecha_inicio_str, fecha_fin_str, None, None, descripcion
                
            except ValueError as e:
                print("[ERROR] Formato de fecha invalido. Use YYYY-MM-DD (ej: 2025-10-01)")
                print("    Error: {}".format(str(e)))
                return None, None, None, None, None, None
        
        # Si no hay rango, usar MODO 1 (mes completo)
        if mes is not None and anio is not None:
            if mes < 1 or mes > 12:
                print("[ERROR] Mes invalido: {}. Debe estar entre 1 y 12".format(mes))
                return None, None, None, None, None, None
            
            if anio < 2020 or anio > 2030:
                print("[ADVERTENCIA] Año inusual: {}".format(anio))
            
            # Calcular primer y ultimo dia del mes
            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
            fecha_fin_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)
            
            # Descripcion para el mes completo
            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)
            
            print("[INFO] Modo: MES COMPLETO")
            return 'mes', fecha_inicio_str, fecha_fin_str, mes, anio, descripcion
        
        # Si no hay ninguno de los dos modos configurados
        print("[ERROR] El archivo {} no contiene configuracion valida".format(config_file))
        print("    Debe tener MES+AÑO o FECHA_INICIO+FECHA_FIN")
        return None, None, None, None, None, None
        
    except Exception as e:
        print("[ERROR] Error leyendo archivo de configuracion: {}".format(str(e)))
        return None, None, None, None, None, None

def get_month_name(mes):
    """Retorna el nombre del mes en español"""
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    """Retorna la abreviatura del mes en español"""
    if mes is None:
        return 'rango'
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_query(fecha_inicio, fecha_fin):
    """Construye la query de Sesiones Abiertas por Pushes con el rango de fechas especificado"""
    
    query = """SELECT starting_cause, count(distinct (session_id)) as Cant_sesiones 
FROM "caba-piba-consume-zone-db"."boti_session_metrics_2"   
WHERE CAST(session_creation_time AS DATE) BETWEEN date '{fecha_inicio}' and date '{fecha_fin}' 
group by starting_cause""".format(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    
    return query

def generate_filename(modo, mes, anio, fecha_inicio, fecha_fin):
    """Genera el nombre del archivo basado en el modo y las fechas"""
    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_csv = "sesiones_abiertas_pushes_{0}_{1}.csv".format(mes_nombre, anio)
        filename_excel = "sesiones_abiertas_pushes_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:  # modo == 'rango'
        fecha_inicio_fmt = fecha_inicio.replace('-', '')
        fecha_fin_fmt = fecha_fin.replace('-', '')
        filename_csv = "sesiones_abiertas_pushes_{0}_a_{1}.csv".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_excel = "sesiones_abiertas_pushes_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
    
    return filename_csv, filename_excel

def create_excel_with_dashboard(filepath, result_value, modo, mes, anio, fecha_inicio, fecha_fin):
    """
    Crea un Excel NUEVO desde cero con estructura de Dashboard completa
    Escribe el resultado SOLO en la celda D4 (Sesiones abiertas por Pushes)
    """
    
    print("    [INFO] Creando Excel NUEVO con estructura Dashboard...")
    
    # IMPORTANTE: Siempre crea un workbook NUEVO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dashboard'
    
    # Formato para encabezados
    header_font = Font(bold=True)
    
    # Determinar el texto del encabezado de fecha
    if modo == 'mes':
        header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])  # Formato: oct-25
    else:  # modo == 'rango'
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        header_fecha = '{}-{}'.format(
            fecha_inicio_obj.strftime('%d/%m'),
            fecha_fin_obj.strftime('%d/%m/%y')
        )
    
    # FILA 1: Encabezados
    ws['B1'] = 'Indicador'
    ws['C1'] = 'Descripción/Detalle'
    ws['D1'] = header_fecha
    ws['B1'].font = header_font
    ws['C1'].font = header_font
    ws['D1'].font = header_font
    
    # FILA 2: Conversaciones
    ws['B2'] = 'Conversaciones'
    ws['C2'] = 'Q Conversaciones'
    # D2 vacío (no se llena)
    
    # FILA 3: Usuarios
    ws['B3'] = 'Usuarios'
    ws['C3'] = 'Q Usuarios únicos'
    # D3 vacío (no se llena)
    
    # FILA 4: Sesiones abiertas por Pushes - AQUI VA EL RESULTADO
    ws['B4'] = 'Sesiones abiertas por Pushes'
    ws['C4'] = 'Q Sesiones que se abrieron con una Push'
    ws['D4'] = result_value  # ← UNICO VALOR QUE SE ESCRIBE
    
    # FILA 5: Sesiones Alcanzadas por Pushes
    ws['B5'] = 'Sesiones Alcanzadas por Pushes'
    ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'
    # D5 vacío (no se llena)
    
    # FILA 6: Mensajes Pushes Enviados
    ws['B6'] = 'Mensajes Pushes Enviados'
    ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'
    # D6 vacío (no se llena)
    
    # FILA 7: Contenidos en Botmaker
    ws['B7'] = 'Contenidos en Botmaker'
    ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'
    # D7 vacío (no se llena)
    
    # FILA 8: Contenidos Prendidos para el USUARIO
    ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
    ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'
    # D8 vacío (no se llena)
    
    # FILA 9: Interacciones
    ws['B9'] = 'Interacciones'
    ws['C9'] = 'Q Interacciones'
    # D9 vacío (no se llena)
    
    # FILA 10: Trámites, solicitudes y turnos
    ws['B10'] = 'Trámites, solicitudes y turnos'
    ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'
    # D10 vacío (no se llena)
    
    # FILA 11: contenidos mas consultados
    ws['B11'] = 'contenidos mas consultados'
    ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'
    # D11 vacío (no se llena)
    
    # FILA 12: Derivaciones
    ws['B12'] = 'Derivaciones'
    ws['C12'] = 'Q Derivaciones'
    # D12 vacío (no se llena)
    
    # FILA 13: No entendimiento
    ws['B13'] = 'No entendimiento'
    ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'
    # D13 vacío (no se llena)
    
    # FILA 14: Tasa de Efectividad
    ws['B14'] = 'Tasa de Efectividad'
    ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo [Estadísticas Eventos]'
    # D14 vacío (no se llena)
    
    # FILA 15: CES (Customer Effort Score)
    ws['B15'] = 'CES (Customer Effort Score)'
    ws['C15'] = 'Puntuación del esfuerzo del cliente [Estadísticas Eventos]'
    # D15 vacío (no se llena)
    
    # Ajustar anchos de columna
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    
    # Guardar
    wb.save(filepath)
    print("    [OK] Excel generado: {}".format(filepath))

def check_aws_credentials():
    """Verifica que las credenciales AWS esten configuradas y sean validas"""
    try:
        sts = boto3.client('sts', region_name=CONFIG['region'])
        identity = sts.get_caller_identity()
        
        user_arn = identity.get('Arn', '')
        
        print("[OK] Credenciales AWS validas")
        print("    ARN: {}".format(user_arn))
        
        # Verificar que sea el rol correcto
        if 'PIBAConsumeBoti' not in user_arn:
            print("")
            print("[ADVERTENCIA] No estas usando el rol correcto")
            print("    Se requiere: PIBAConsumeBoti")
            if '/' in user_arn:
                current_role = user_arn.split('/')[-2]
            else:
                current_role = 'desconocido'
            print("    Tu rol actual: {}".format(current_role))
            print("")
            print("SOLUCION:")
            print("    1. Ejecuta: aws-azure-login --profile default --mode=gui")
            print("    2. Cuando te autentiques, SELECCIONA el rol: PIBAConsumeBoti")
            print("    3. Vuelve a ejecutar este script")
            print("")
            return False
        
        return True
        
    except Exception as e:
        print("[ERROR] Error verificando credenciales: {}".format(str(e)))
        if 'ExpiredToken' in str(e):
            print("")
            print("SOLUCION:")
            print("    Tu sesión AWS expiró. Ejecuta:")
            print("    aws-azure-login --profile default --mode=gui")
            print("")
        else:
            print("")
            print("SOLUCION:")
            print("    1. Ejecuta: aws-azure-login --configure --profile default")
            print("    2. Luego: aws-azure-login --profile default --mode=gui")
            print("")
        return False

def execute_query_and_save():
    """Funcion principal: ejecuta query y guarda resultados"""
    
    # Verificar credenciales
    print("Verificando credenciales AWS...")
    if not check_aws_credentials():
        return None
    
    # Leer configuracion de fechas
    print("")
    print("Leyendo configuracion de fechas...")
    
    result = read_date_config(CONFIG['config_file'])
    
    if result[0] is None:
        print("[ERROR] No se pudo leer la configuracion de fechas")
        return None
    
    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = result
    
    print("[OK] Configuracion leida:")
    print("    Periodo: {}".format(descripcion))
    print("    Fecha inicio: {}".format(fecha_inicio))
    print("    Fecha fin: {}".format(fecha_fin))
    
    # Construir query
    query = build_query(fecha_inicio, fecha_fin)
    
    print("")
    print("Configuracion AWS:")
    print("    Region: {}".format(CONFIG['region']))
    print("    Workgroup: {}".format(CONFIG['workgroup']))
    print("    Base de datos: {}".format(CONFIG['database']))
    
    print("")
    print("Query a ejecutar:")
    print("    {}".format(query))
    
    try:
        # Crear sesion boto3
        session = boto3.Session(region_name=CONFIG['region'])
        
        print("")
        print("Ejecutando consulta...")
        
        # Intentar con el workgroup especificado
        try:
            df = wr.athena.read_sql_query(
                sql=query,
                database=CONFIG['database'],
                workgroup=CONFIG['workgroup'],
                boto3_session=session,
                ctas_approach=False,
                unload_approach=False
            )
        except Exception as e:
            if 'workgroup' in str(e).lower() or 'GetWorkGroup' in str(e):
                print("")
                print("[ADVERTENCIA] Error con workgroup '{}'".format(CONFIG['workgroup']))
                print("    Intentando sin especificar workgroup...")
                
                df = wr.athena.read_sql_query(
                    sql=query,
                    database=CONFIG['database'],
                    boto3_session=session,
                    ctas_approach=False,
                    unload_approach=False
                )
            else:
                raise e
        
        print("")
        print("[OK] Consulta ejecutada exitosamente!")
        
        # Procesar resultados (puede haber múltiples filas por el GROUP BY)
        if len(df) > 0 and 'starting_cause' in df.columns and 'Cant_sesiones' in df.columns:
            # Buscar el valor correspondiente a 'WhatsAppTemplate' (sesiones abiertas por pushes)
            whatsapp_row = df[df['starting_cause'] == 'WhatsAppTemplate']
            
            if len(whatsapp_row) > 0:
                result_value = int(whatsapp_row['Cant_sesiones'].iloc[0])
            else:
                print("[ADVERTENCIA] No se encontró 'WhatsAppTemplate' en starting_cause")
                print("    Valores encontrados: {}".format(df['starting_cause'].tolist()))
                # Si no hay 'WhatsAppTemplate', usar 0
                result_value = 0
        else:
            print("[ERROR] No se pudo obtener el resultado de la query")
            print("    Columnas: {}".format(df.columns.tolist() if len(df) > 0 else 'Sin datos'))
            return None
        
        # Mostrar resultados detallados
        print("")
        print("=" * 60)
        print("RESULTADOS - {}".format(descripcion.upper()))
        print("=" * 60)
        print("\nDesglose por starting_cause:")
        for idx, row in df.iterrows():
            print("  {}: {:,}".format(row['starting_cause'], row['Cant_sesiones']))
        
        print("\n" + "=" * 60)
        print("SESIONES ABIERTAS POR PUSHES (WhatsAppTemplate): {:,}".format(result_value))
        print("=" * 60)
        
        # Generar nombres de archivo
        filename_csv, filename_excel = generate_filename(modo, mes, anio, fecha_inicio, fecha_fin)
        output_folder = CONFIG['output_folder']
        
        # Crear carpeta si no existe
        os.makedirs(output_folder, exist_ok=True)
        
        # Rutas completas
        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_excel = os.path.join(output_folder, filename_excel)
        
        # Guardar CSV
        print("")
        print("Guardando CSV...")
        df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        
        # Crear Excel con Dashboard y resultado en D4
        print("Generando Excel Dashboard...")
        create_excel_with_dashboard(local_path_excel, result_value, modo, mes, anio, fecha_inicio, fecha_fin)
        
        print("")
        print("ARCHIVOS GENERADOS:")
        print("    Carpeta: {}/".format(output_folder))
        print("")
        print("    [CSV] Nombre: {}".format(filename_csv))
        print("          Ruta: {}".format(os.path.abspath(local_path_csv)))
        print("          Tamaño: {:,} bytes".format(os.path.getsize(local_path_csv)))
        print("")
        print("    [EXCEL] Nombre: {}".format(filename_excel))
        print("            Ruta: {}".format(os.path.abspath(local_path_excel)))
        print("            Tamaño: {:,} bytes".format(os.path.getsize(local_path_excel)))
        print("            Hoja: Dashboard")
        print("            Resultado en celda: D4 = {:,}".format(result_value))
        print("            [IMPORTANTE] Excel creado NUEVO con estructura completa")
        
        print("")
        print("=" * 60)
        print("PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 60)
        
        return df
        
    except Exception as e:
        print("")
        print("[ERROR] ERROR DURANTE LA EJECUCION")
        print("    Tipo: {}".format(type(e).__name__))
        print("    Mensaje: {}".format(str(e)))
        
        error_str = str(e).lower()
        
        print("")
        print("DIAGNOSTICO:")
        if 'table' in error_str and 'not' in error_str:
            print("    [!] La tabla no existe o no tienes permisos para accederla")
            print("    Verifica acceso a: boti_session_metrics_2")
        elif 'workgroup' in error_str:
            print("    [!] Problema con el workgroup")
        elif 'permission' in error_str or 'denied' in error_str:
            print("    [!] Problema de permisos")
        elif 'openpyxl' in error_str:
            print("    [!] Falta libreria openpyxl para generar Excel")
            print("    Ejecuta: pip install openpyxl")
        elif 'timeout' in error_str or 'timed out' in error_str:
            print("    [!] La query tomó demasiado tiempo")
        else:
            print("    [!] Error inesperado")
        
        return None

# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    print("")
    print("=" * 60)
    print("SCRIPT: SESIONES ABIERTAS POR PUSHES - QUERY ATHENA V2")
    print("=" * 60)
    print("Lee configuracion desde: {}".format(CONFIG['config_file']))
    print("Rol requerido: PIBAConsumeBoti")
    print("Salida: CSV + Excel Dashboard NUEVO (resultado en celda D4)")
    print("Query: boti_session_metrics_2 agrupado por starting_cause")
    print("")
    print("MODOS SOPORTADOS:")
    print("  [1] MES COMPLETO: Configura MES y AÑO")
    print("  [2] RANGO PERSONALIZADO: Configura FECHA_INICIO y FECHA_FIN")
    print("=" * 60)
    print("")
    
    result = execute_query_and_save()
    
    if result is not None:
        print("")
        print("Listo! Tus archivos estan guardados.")
        print("")
        print("Para procesar otro periodo:")
        print("    1. Edita el archivo: {}".format(CONFIG['config_file']))
        print("    2. Cambia las fechas segun el modo deseado")
        print("    3. Vuelve a ejecutar este script")
    else:
        print("")
        print("La ejecucion fallo. Revisa los mensajes de error arriba.")
