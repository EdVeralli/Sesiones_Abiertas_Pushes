# -*- coding: utf-8 -*-
"""

aws-azure-login --configure --profile default

aws-azure-login --profile default --mode=gui



Script para ejecutar query de Mensajes Pushes Enviados en Athena
Guarda resultado en CSV y escribe el resultado en la celda D6 de Excel
Lee configuracion de fechas desde archivo config_fechas.txt
IMPORTANTE: El Excel SIEMPRE se crea NUEVO desde cero con estructura de Dashboard
Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBAConsumeBoti
"""
import boto3
import awswrangler as wr
import pandas as pd
from datetime import datetime
from calendar import monthrange
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': 'config_fechas.txt'
}

# ==================== FUNCIONES ====================

def read_date_config(config_file):
    """Lee el archivo de configuracion y extrae MES y AÑO"""
    try:
        if not os.path.exists(config_file):
            print("[ERROR] No se encuentra el archivo: {}".format(config_file))
            print("    Creando archivo de ejemplo...")
            
            with open(config_file, 'w', encoding='utf-8') as f:
                f.write("# Configuracion de fecha para filtro automatico\n")
                f.write("# Formato: MES=numero del mes (1-12)\n")
                f.write("# Formato: AÑO=año completo (ej: 2025)\n\n")
                f.write("MES=9\n")
                f.write("AÑO=2024\n")
            
            print("    Archivo creado: {}".format(config_file))
            return 9, 2024
        
        mes = None
        anio = None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)
                
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)
        
        if mes is None or anio is None:
            print("[ERROR] El archivo {} no contiene MES o AÑO validos".format(config_file))
            return None, None
        
        if mes < 1 or mes > 12:
            print("[ERROR] Mes invalido: {}. Debe estar entre 1 y 12".format(mes))
            return None, None
        
        if anio < 2020 or anio > 2030:
            print("[ADVERTENCIA] Año inusual: {}".format(anio))
        
        return mes, anio
        
    except Exception as e:
        print("[ERROR] Error leyendo archivo de configuracion: {}".format(str(e)))
        return None, None

def get_month_name(mes):
    """Retorna el nombre del mes en español"""
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    """Retorna la abreviatura del mes en español"""
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_query(mes, anio):
    """Construye la query de Mensajes Pushes Enviados con el mes y año especificados"""
    
    # Calcular primer y último día del mes
    primer_dia = 1
    ultimo_dia = monthrange(anio, mes)[1]
    
    fecha_inicio = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
    fecha_fin = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)
    
    query = """SELECT count(distinct m.id) as count_messages
FROM "caba-piba-consume-zone-db"."boti_event_metrics_2" ev 
JOIN "caba-piba-consume-zone-db"."boti_message_metrics_2" m 
ON ev.session_id=m.session_id 
WHERE CAST(ev.creation_time AS DATE) BETWEEN date '{fecha_inicio}' and date '{fecha_fin}'  
AND regexp_like(m.message, '^Template') 
and events_name in ('notification-status-sent')""".format(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    
    return query

def generate_filename(mes, anio):
    """Genera el nombre del archivo basado en mes y año"""
    mes_nombre = get_month_name(mes)
    filename_csv = "mensajes_pushes_enviados_{0}_{1}.csv".format(mes_nombre, anio)
    filename_excel = "mensajes_pushes_enviados_{0}_{1}.xlsx".format(mes_nombre, anio)
    return filename_csv, filename_excel

def create_excel_with_dashboard(filepath, result_value, mes, anio):
    """
    Crea un Excel NUEVO desde cero con estructura de Dashboard completa
    Escribe el resultado SOLO en la celda D6 (Mensajes Pushes Enviados)
    """
    
    print("    [INFO] Creando Excel NUEVO con estructura Dashboard...")
    
    # IMPORTANTE: Siempre crea un workbook NUEVO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dashboard'
    
    # Formato para encabezados
    header_font = Font(bold=True)
    
    # FILA 1: Encabezados
    ws['B1'] = 'Indicador'
    ws['C1'] = 'Descripción/Detalle'
    ws['D1'] = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])  # Formato: sep-24
    ws['B1'].font = header_font
    ws['C1'].font = header_font
    ws['D1'].font = header_font
    
    # FILA 2: Conversaciones
    ws['B2'] = 'Conversaciones'
    ws['C2'] = 'Q Conversaciones'
    # D2 vacío (no se llena)
    
    # FILA 3: Usuarios
    ws['B3'] = 'Usuarios'
    ws['C3'] = 'Q Usuarios únicos'
    # D3 vacío (no se llena)
    
    # FILA 4: Sesiones abiertas por Pushes
    ws['B4'] = 'Sesiones abiertas por Pushes'
    ws['C4'] = 'Q Sesiones que se abrieron con una Push'
    # D4 vacío (no se llena)
    
    # FILA 5: Sesiones Alcanzadas por Pushes
    ws['B5'] = 'Sesiones Alcanzadas por Pushes'
    ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'
    # D5 vacío (no se llena)
    
    # FILA 6: Mensajes Pushes Enviados - AQUI VA EL RESULTADO
    ws['B6'] = 'Mensajes Pushes Enviados'
    ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'
    ws['D6'] = result_value  # ← UNICO VALOR QUE SE ESCRIBE
    
    # FILA 7: Contenidos en Botmaker
    ws['B7'] = 'Contenidos en Botmaker'
    ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'
    # D7 vacío (no se llena)
    
    # FILA 8: Contenidos Prendidos para el USUARIO
    ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
    ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'
    # D8 vacío (no se llena)
    
    # FILA 9: Interacciones
    ws['B9'] = 'Interacciones'
    ws['C9'] = 'Q Interacciones'
    # D9 vacío (no se llena)
    
    # FILA 10: Trámites, solicitudes y turnos
    ws['B10'] = 'Trámites, solicitudes y turnos'
    ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'
    # D10 vacío (no se llena)
    
    # FILA 11: contenidos mas consultados
    ws['B11'] = 'contenidos mas consultados'
    ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'
    # D11 vacío (no se llena)
    
    # FILA 12: Derivaciones
    ws['B12'] = 'Derivaciones'
    ws['C12'] = 'Q Derivaciones'
    # D12 vacío (no se llena)
    
    # FILA 13: No entendimiento
    ws['B13'] = 'No entendimiento'
    ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'
    # D13 vacío (no se llena)
    
    # FILA 14: Tasa de Efectividad
    ws['B14'] = 'Tasa de Efectividad'
    ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo [Estadísticas Eventos]'
    # D14 vacío (no se llena)
    
    # FILA 15: CES (Customer Effort Score)
    ws['B15'] = 'CES (Customer Effort Score)'
    ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar con'
    # D15 vacío (no se llena)
    
    # FILA 16: Satisfacción (CSAT)
    ws['B16'] = 'Satisfacción (CSAT)'
    ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5, donde 1 es "muy insatisfecho"'
    # D16 vacío (no se llena)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    
    # Guardar el archivo
    wb.save(filepath)
    print("    [OK] Excel creado en: {}".format(filepath))

def check_aws_credentials():
    """Verifica que las credenciales AWS esten configuradas y sean validas"""
    try:
        sts = boto3.client('sts', region_name=CONFIG['region'])
        identity = sts.get_caller_identity()
        user_arn = identity['Arn']
        
        print("[OK] Credenciales AWS encontradas")
        print("    Usuario: {}".format(user_arn))
        
        if 'PIBAConsumeBoti' in user_arn:
            print("[OK] Rol correcto: PIBAConsumeBoti")
        else:
            print("[ADVERTENCIA] Rol actual no es PIBAConsumeBoti")
            print("    Se requiere: PIBAConsumeBoti")
            if '/' in user_arn:
                current_role = user_arn.split('/')[-2]
            else:
                current_role = 'desconocido'
            print("    Tu rol actual: {}".format(current_role))
            print("")
            print("SOLUCION:")
            print("    1. Ejecuta: aws-azure-login --profile default --mode=gui")
            print("    2. Cuando te autentiques, SELECCIONA el rol: PIBAConsumeBoti")
            print("    3. Vuelve a ejecutar este script")
            print("")
            return False
        
        return True
        
    except Exception as e:
        print("[ERROR] Error verificando credenciales: {}".format(str(e)))
        if 'ExpiredToken' in str(e):
            print("")
            print("SOLUCION:")
            print("    Tu sesión AWS expiró. Ejecuta:")
            print("    aws-azure-login --profile default --mode=gui")
            print("")
        else:
            print("")
            print("SOLUCION:")
            print("    1. Ejecuta: aws-azure-login --configure --profile default")
            print("    2. Luego: aws-azure-login --profile default --mode=gui")
            print("")
        return False

def execute_query_and_save():
    """Funcion principal: ejecuta query y guarda resultados"""
    
    # Verificar credenciales
    print("Verificando credenciales AWS...")
    if not check_aws_credentials():
        return None
    
    # Leer configuracion de fechas
    print("")
    print("Leyendo configuracion de fechas...")
    
    mes, anio = read_date_config(CONFIG['config_file'])
    
    if mes is None or anio is None:
        print("[ERROR] No se pudo leer la configuracion de fechas")
        return None
    
    mes_nombre = get_month_name(mes)
    
    print("[OK] Configuracion leida:")
    print("    Mes: {} ({})".format(mes, mes_nombre.capitalize()))
    print("    Año: {}".format(anio))
    
    # Construir query dinamicamente
    query = build_query(mes, anio)
    
    print("")
    print("Configuracion AWS:")
    print("    Region: {}".format(CONFIG['region']))
    print("    Workgroup: {}".format(CONFIG['workgroup']))
    print("    Base de datos: {}".format(CONFIG['database']))
    
    print("")
    print("Query a ejecutar:")
    print("    {}".format(query))
    
    try:
        # Crear sesion boto3
        session = boto3.Session(region_name=CONFIG['region'])
        
        print("")
        print("Ejecutando consulta...")
        print("[INFO] Esta query puede tardar debido al JOIN entre tablas...")
        
        # Intentar con el workgroup especificado
        try:
            df = wr.athena.read_sql_query(
                sql=query,
                database=CONFIG['database'],
                workgroup=CONFIG['workgroup'],
                boto3_session=session,
                ctas_approach=False,
                unload_approach=False
            )
        except Exception as e:
            if 'workgroup' in str(e).lower() or 'GetWorkGroup' in str(e):
                print("")
                print("[ADVERTENCIA] Error con workgroup '{}'".format(CONFIG['workgroup']))
                print("    Intentando sin especificar workgroup...")
                
                df = wr.athena.read_sql_query(
                    sql=query,
                    database=CONFIG['database'],
                    boto3_session=session,
                    ctas_approach=False,
                    unload_approach=False
                )
            else:
                raise e
        
        print("")
        print("[OK] Consulta ejecutada exitosamente!")
        
        # Extraer el resultado (deberia ser un solo valor)
        if len(df) > 0 and 'count_messages' in df.columns:
            result_value = int(df['count_messages'].iloc[0])
        else:
            print("[ERROR] No se pudo obtener el resultado de la query")
            print("    Columnas: {}".format(df.columns.tolist() if len(df) > 0 else 'Sin datos'))
            return None
        
        # Mostrar resultado
        print("")
        print("=" * 60)
        print("RESULTADO - {} {}".format(mes_nombre.upper(), anio))
        print("=" * 60)
        print("Mensajes Pushes Enviados: {:,}".format(result_value))
        print("=" * 60)
        
        # Generar nombres de archivo
        filename_csv, filename_excel = generate_filename(mes, anio)
        output_folder = CONFIG['output_folder']
        
        # Crear carpeta si no existe
        os.makedirs(output_folder, exist_ok=True)
        
        # Rutas completas
        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_excel = os.path.join(output_folder, filename_excel)
        
        # Guardar CSV
        print("")
        print("Guardando CSV...")
        df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        
        # Crear Excel con Dashboard y resultado en D6
        print("Generando Excel Dashboard...")
        create_excel_with_dashboard(local_path_excel, result_value, mes, anio)
        
        print("")
        print("ARCHIVOS GENERADOS:")
        print("    Carpeta: {}/".format(output_folder))
        print("")
        print("    [CSV] Nombre: {}".format(filename_csv))
        print("          Ruta: {}".format(os.path.abspath(local_path_csv)))
        print("          Tamaño: {:,} bytes".format(os.path.getsize(local_path_csv)))
        print("")
        print("    [EXCEL] Nombre: {}".format(filename_excel))
        print("            Ruta: {}".format(os.path.abspath(local_path_excel)))
        print("            Tamaño: {:,} bytes".format(os.path.getsize(local_path_excel)))
        print("            Hoja: Dashboard")
        print("            Resultado en celda: D6 = {:,}".format(result_value))
        print("            [IMPORTANTE] Excel creado NUEVO con estructura completa")
        
        print("")
        print("=" * 60)
        print("PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 60)
        
        return df
        
    except Exception as e:
        print("")
        print("[ERROR] ERROR DURANTE LA EJECUCION")
        print("    Tipo: {}".format(type(e).__name__))
        print("    Mensaje: {}".format(str(e)))
        
        error_str = str(e).lower()
        
        print("")
        print("DIAGNOSTICO:")
        if 'table' in error_str and 'not' in error_str:
            print("    [!] La tabla no existe o no tienes permisos para accederla")
            print("    Verifica acceso a: boti_event_metrics_2 y boti_message_metrics_2")
        elif 'workgroup' in error_str:
            print("    [!] Problema con el workgroup")
        elif 'permission' in error_str or 'denied' in error_str:
            print("    [!] Problema de permisos")
        elif 'openpyxl' in error_str:
            print("    [!] Falta libreria openpyxl para generar Excel")
            print("    Ejecuta: pip install openpyxl")
        elif 'timeout' in error_str or 'timed out' in error_str:
            print("    [!] La query tomó demasiado tiempo (JOIN puede ser costoso)")
        else:
            print("    [!] Error inesperado")
        
        return None

# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    print("")
    print("=" * 60)
    print("SCRIPT: MENSAJES PUSHES ENVIADOS - QUERY ATHENA")
    print("=" * 60)
    print("Lee configuracion desde: {}".format(CONFIG['config_file']))
    print("Rol requerido: PIBAConsumeBoti")
    print("Salida: CSV + Excel Dashboard NUEVO (resultado en celda D6)")
    print("Query: JOIN entre boti_event_metrics_2 y boti_message_metrics_2")
    print("=" * 60)
    print("")
    
    result = execute_query_and_save()
    
    if result is not None:
        print("")
        print("Listo! Tus archivos estan guardados.")
        print("")
        print("Para procesar otro mes/año:")
        print("    1. Edita el archivo: {}".format(CONFIG['config_file']))
        print("    2. Cambia MES y AÑO")
        print("    3. Vuelve a ejecutar este script")
    else:
        print("")
        print("La ejecucion fallo. Revisa los mensajes de error arriba.")
